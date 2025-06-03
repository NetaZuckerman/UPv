#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Mar 10 05:42:03 2024

@author: hagar
"""

import pandas as pd
from pipelines.generalPipeline import general_pipe
import os 
from utils.utils import get_sequences, write_sub_fasta, create_dirs
from mutations import signatures
from utils.format_xl import save_format_xl
import pysam
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
SCRIPT_PATH = os.path.dirname(__file__) +'/'



ul54_reg = (77858, 81587) 	
ul97_reg = (141440, 143564)


class cmv(general_pipe):
    def __init__(self, reference, fastq, minion, threads):
        super().__init__(reference, fastq, minion, threads)    
        self.reference = SCRIPT_PATH + "refs/CMV_FJ527563.1.fasta"
        
    
    #cut by gene
    def cut_genes(self, aln_path):
        '''
        cut the genes from the all_aligned file.

        Parameters
        ----------
        aln_path : str
            path to alignment fasta file.

        Returns
        -------
        None.

        '''
        fasta = get_sequences(aln_path + "all_aligned.fasta")
        write_sub_fasta(fasta, aln_path, ul97_reg, "reg_UL97")
        write_sub_fasta(fasta, aln_path, ul54_reg, "reg_UL54", '-')
        
            
    #mutations
    def mutations(self):
        '''
        run mutation analysis for each gene seperatly.

        Returns
        -------
        None.

        '''
        create_dirs(["reports"]) 
        signatures.run("alignment/reg_UL54.fasta", "", "reports/mutations_UL54.xlsx")
        signatures.run("alignment/reg_UL97.fasta", "", "reports/mutations_UL97.xlsx")
        

    def split_deletion(self, row):
        aa_change = row['aa_change']
        if 'del' in aa_change:
            positions = aa_change[3:].split('-')
            
            # Handle cases like "del413" by converting to "413-"
            if len(positions) == 1:
                return pd.DataFrame({**row.to_dict(), 'aa_change': [f"{positions[0]}-"]})
            
            # Handle cases like "del413-414"
            if len(positions) == 2:
                start, end = map(int, positions)
                return pd.DataFrame({**row.to_dict(), 'aa_change': [f"{pos}-" for pos in range(start, end + 1)]})
        
        # Return the entire row unchanged for non-deletions
        return pd.DataFrame([row])


    def resist_poly_reports(self):
        '''
        use the mutations table to generate resistant and polymorphism reports

        Returns
        -------
        None.

        '''
        
        # load mutations database
        resist = pd.read_csv(SCRIPT_PATH + "refs/herpesdrg-db.tsv", sep='\t').rename(columns=({"gene": "gene_name"}))
        resist = resist[resist["virus"] == 'HCMV']
        resist = pd.concat(resist.apply(lambda row: self.split_deletion(row), axis=1).to_list(), ignore_index=True)

        resist['aa_position_on_gene'] = resist['aa_change'].apply(lambda x: x[1:-1] if '-' not in x else x[:-1])
        
        #open mutation tables for each gene
        mut_tbl_UL54 = pd.read_excel("reports/mutations_UL54.xlsx")
        mut_tbl_UL54["gene_name"] = "UL54"
        mut_tbl_UL54["nt_position_on_genome"] = mut_tbl_UL54["nt_position_on_genome"] + ul54_reg[0]-1
        mut_tbl_UL97 = pd.read_excel("reports/mutations_UL97.xlsx")
        mut_tbl_UL97["gene_name"] = "UL97"
        mut_tbl_UL97["nt_position_on_genome"] = mut_tbl_UL97["nt_position_on_genome"] + ul97_reg[0]-1
        
        #concat all gene to one table
        mut_tbl = pd.concat([mut_tbl_UL54,mut_tbl_UL97])
        mut_tbl['aa_position_on_gene'] = mut_tbl['aa_position_on_gene'].astype(str)
        # merge mutation table with the herpesdrg database
        merged = mut_tbl.merge(resist, how='left', on=["gene_name","aa_position_on_gene"])
        
        #save the mutation table in excel
        seq_num = len(self.sample_fq_dict)
        save_format_xl(merged, seq_num, "reports/mutations.xlsx")

        #remove script temp files:
        # os.remove("reports/mutations_UL54.xlsx")
        # os.remove("reports/mutations_UL97.xlsx")


    def excel_colors(self, df):
        # Save to Excel first
        excel_path = 'reports/summary.xlsx'
        df.to_excel(excel_path, index=False)
        
        # Now open it with openpyxl and color rows
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Color setup
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Track and alternate color per sample
        current_sample = None
        use_gray = False
        for row_idx in range(2, ws.max_row + 1):  # start from row 2 (row 1 is header)
            sample_value = ws.cell(row=row_idx, column=1).value
            if sample_value != current_sample:
                use_gray = not use_gray
                current_sample = sample_value
        
            fill = gray_fill if use_gray else white_fill
        
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        
        # Save the modified Excel
        wb.save(excel_path)
    
    
    def resist_poly_summary(self):
        samples = list(self.sample_fq_dict.keys())
        drugs = ['Ganciclovir','Aciclovir', 'Cidofovir', 'Foscarnet', 'Brincidofovir',
                 'Letermovir','Brivudine', 'Penciclovir', 'Tomeglovir',
                 'Maribavir', 'Cyclopropavir','Amenamevir']

        mut_tbl = pd.read_excel("reports/mutations.xlsx")
        mut_tbl = mut_tbl[mut_tbl["R/S"] == 'R']
        mut_tbl = mut_tbl[mut_tbl["mutation_id"].notna()]



        drugs_sum = pd.DataFrame(columns=['sample', 'gene', 'mutation', 'drugs', 'notes'])
        for index, row in mut_tbl.iterrows():
            gene = row['gene_name']
            for sample in samples:
                if not row[sample + "_AA"] == row["FJ527563.1_AA"]: #check if the sample has a mutation in this position
                    mutation = row["aa_change"]
                    if mutation[-1] == row[sample + "_AA"]: #check if the mutation is same as in database
                        relevant_drugs = [drug for drug in drugs if not str(row[drug]) == 'nan' and not str(row[drug]) == ' ']
                        for drug in relevant_drugs:    
                            drugs_sum.loc[len(drugs_sum)] = [sample, gene, mutation, drug, row[drug]]

        # poly_df = drugs_sum[drugs_sum["notes"] == "Polymorphism"].drop(columns=["drugs","notes"]).drop_duplicates().sort_values(by=['sample'])
        # resist_df = drugs_sum[drugs_sum["notes"] != "Polymorphism"].drop_duplicates().sort_values(by=['sample'])
        
        # poly_df.to_excel("reports/polymorphism_summary.xlsx", index=False)
        # resist_df.to_excel("reports/resist_summary.xlsx", index=False)
        drugs_sum.to_excel("reports/summary.xlsx", index=False)
        drugs_sum = drugs_sum.drop_duplicates().sort_values(by=['sample','notes'])
        self.excel_colors(drugs_sum)
        
        
    #override report
    def qc_report(self, bam_path, depth_path, output_report):
        '''
        run unique hsv functions.        
        generate qc report considering the genes.
        Parameters
        ----------
        bam_path : str
            path to bam folder.
        depth_path : str
            path to depth folder.
        output_report : str
            path to qc report file.

        Returns
        -------
        None.

        '''
        self.cut_genes("alignment/")
        self.mutations()
        self.resist_poly_reports()
        self.resist_poly_summary()
        
        general_qc = pd.DataFrame(columns=['sample', 'mapped%','mapped_reads','total_reads','cov_bases',"chimeric_read_count"])
        ul54_qc = pd.DataFrame(columns=['sample','coverage%','coverage_CNS_5%', 'mean_depth','max_depth','min_depth'])
        ul97_qc = pd.DataFrame(columns=['sample','coverage%','coverage_CNS_5%', 'mean_depth','max_depth','min_depth'])
        
        for bam_file in os.listdir(bam_path):
                if "sorted" in bam_file and "bai" not in bam_file:
                    #general qc
                    sample = bam_file.split(".mapped")[0] + bam_file.split(".sorted")[1].split(".bam")[0]
                    total_reads = pysam.AlignmentFile(bam_path + bam_file.split(".mapped")[0]+".bam").count(until_eof=True)
                    mapped_reads, mapped_percentage, cov_bases, chimer_count = super().general_qc(bam_path+bam_file, total_reads)
            
                    #depth 
                    max_depth54, min_depth54,mean_depth54, cover54, cns5_cover54 = super().depth_qc(depth_path+sample+".txt", start = ul54_reg[0], end= ul54_reg[1])
                    max_depth97, min_depth97, mean_depth97, cover97, cns5_cover97 = super().depth_qc(depth_path+sample+".txt", start = ul97_reg[0], end= ul97_reg[1])
                    
                    #add rows
                    general_qc.loc[len(general_qc.index)] = [sample, mapped_percentage, mapped_reads, total_reads, cov_bases, chimer_count]
                    ul54_qc.loc[len(ul54_qc.index)] = [sample, cover54, cns5_cover54, mean_depth54, max_depth54, min_depth54]
                    ul97_qc.loc[len(ul97_qc.index)] = [sample, cover97, cns5_cover97, mean_depth97, max_depth97, min_depth97]
                    
        with pd.ExcelWriter('QC/QC_report.xlsx') as writer:  
            general_qc.to_excel(writer, sheet_name='general', index=False)
            ul54_qc.to_excel(writer, sheet_name='UL54', index=False)
            ul97_qc.to_excel(writer, sheet_name='UL97', index=False)
