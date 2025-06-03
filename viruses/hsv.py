#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct  9 08:05:15 2023

@author: hagar


hsv pipeline was generated to analysis Human Herpes Virus 1. 
The main purpose is to generate fasta sequence for each gene and to find resistance mutations and polymorphysm.
"""

#regions of JN555585
ul30_reg = (62807, 66515) 	
ul42_reg = (93112, 94579)
ul23_reg = (46672, 47803)


import pandas as pd
from pipelines.generalPipeline import general_pipe
import os 
from utils.utils import get_sequences, write_sub_fasta, create_dirs
from mutations import signatures
from utils.format_xl import save_format_xl
import pysam
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SCRIPT_PATH = os.path.dirname(__file__) +'/'

class hsv(general_pipe):
    def __init__(self, reference, fastq, minion, threads):
        '''

        Parameters
        ----------
        reference : str
            path to the reference fasta file. no need to provide reference.
        fastq : str
            path to fastq folder
        minion : BOOL
            boolean to indicate if the the reads are minion based. defualt is illumina
        threads : int
            max number of threads for parts threads are available in this pipeline.
        

        Returns
        -------
        None.

        '''
        super().__init__(reference, fastq, minion, threads)    
        self.reference = SCRIPT_PATH + "refs/JN555585.fasta"

#cut by gene
    def cut_genes(self, aln_path):
        '''
        cut the genes from the all_aligned file.

        Parameters
        ----------
        aln_path : str
            path to alignment fasta file.

        Returns
        -------
        None.

        '''
        fasta = get_sequences(aln_path + "all_aligned.fasta")
        write_sub_fasta(fasta, aln_path, ul30_reg, "reg_UL30")
        write_sub_fasta(fasta, aln_path, ul42_reg, "reg_UL42")
        write_sub_fasta(fasta, aln_path, ul23_reg, "reg_UL23", '-')
        
#mutations
    def mutations(self):
        '''
        run mutation analysis for each gene seperatly.

        Returns
        -------
        None.

        '''
        create_dirs(["reports"]) 
        signatures.run("alignment/reg_UL23.fasta", "", "reports/mutations_UL23.xlsx")
        signatures.run("alignment/reg_UL30.fasta", "", "reports/mutations_UL30.xlsx")
        signatures.run("alignment/reg_UL42.fasta", "", "reports/mutations_UL42.xlsx")


    def resist_poly_reports(self):
        '''
        use the mutations table to generate resistant and polymorphism reports

        Returns
        -------
        None.

        '''
        
        # load mutations database
        resist = pd.read_csv(SCRIPT_PATH + "refs/herpesdrg-db.tsv", sep='\t').rename(columns=({"gene": "gene_name"}))
        resist = resist[resist["virus"] == 'HSV1']
        resist = resist[~resist['aa_change'].str.contains('del')] #temp
        resist['aa_position_on_gene'] = resist['aa_change'].str[1:-1].astype(str)
        
        #open mutation tables for each gene
        mut_tbl_UL23 = pd.read_excel("reports/mutations_UL23.xlsx")
        mut_tbl_UL23["gene_name"] = "UL23"
        mut_tbl_UL23["nt_position_on_genome"] = mut_tbl_UL23["nt_position_on_genome"] + ul23_reg[0]-1
        mut_tbl_UL30 = pd.read_excel("reports/mutations_UL30.xlsx")
        mut_tbl_UL30["gene_name"] = "UL30"
        mut_tbl_UL30["nt_position_on_genome"] = mut_tbl_UL30["nt_position_on_genome"] + ul30_reg[0]-1
        mut_tbl_UL42 = pd.read_excel("reports/mutations_UL42.xlsx")
        mut_tbl_UL42["gene_name"] = "UL42"
        mut_tbl_UL42["nt_position_on_genome"] = mut_tbl_UL42["nt_position_on_genome"] + ul42_reg[0]-1
        #concat all gene to one table
        mut_tbl = pd.concat([mut_tbl_UL23,mut_tbl_UL30,mut_tbl_UL42])
        mut_tbl['aa_position_on_gene'] = mut_tbl['aa_position_on_gene'].astype(str)
        # merge mutation table with the herpesdrg database
        merged = mut_tbl.merge(resist, how='left', on=["gene_name","aa_position_on_gene"])
        
        #save the mutation table in excel
        seq_num = len(self.sample_fq_dict)
        save_format_xl(merged, seq_num, "reports/mutations.xlsx")

        #remove script temp files:
        os.remove("reports/mutations_UL23.xlsx")
        os.remove("reports/mutations_UL30.xlsx")
        os.remove("reports/mutations_UL42.xlsx")            
    
    def excel_colors(self, df):
        # Save to Excel first
        excel_path = 'reports/summary.xlsx'
        df.to_excel(excel_path, index=False)
        
        # Now open it with openpyxl and color rows
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Color setup
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Track and alternate color per sample
        current_sample = None
        use_gray = False
        for row_idx in range(2, ws.max_row + 1):  # start from row 2 (row 1 is header)
            sample_value = ws.cell(row=row_idx, column=1).value
            if sample_value != current_sample:
                use_gray = not use_gray
                current_sample = sample_value
        
            fill = gray_fill if use_gray else white_fill
        
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        
        # Save the modified Excel
        wb.save(excel_path)
    
    def resist_poly_summary(self):
        samples = list(self.sample_fq_dict.keys())
        drugs = ['Ganciclovir','Aciclovir', 'Cidofovir', 'Foscarnet', 'Brincidofovir',
                 'Letermovir','Brivudine', 'Penciclovir', 'Tomeglovir',
                 'Maribavir', 'Cyclopropavir','Amenamevir']

        mut_tbl = pd.read_excel("reports/mutations.xlsx")
        mut_tbl = mut_tbl[mut_tbl["R/S"] == 'R']
        mut_tbl = mut_tbl[mut_tbl["mutation_id"].notna()]



        drugs_sum = pd.DataFrame(columns=['sample', 'gene', 'mutation', 'drugs', 'notes'])
        for index, row in mut_tbl.iterrows():
            gene = row['gene_name']
            for sample in samples:
                if not row[sample + "_AA"] == row["JN555585.1_AA"]: #check if the sample has a mutation in this position
                    mutation = row["aa_change"]
                    if mutation[-1] == row[sample + "_AA"]: #check if the mutation is same as in database
                        relevant_drugs = [drug for drug in drugs if not str(row[drug]) == 'nan' and not str(row[drug]) == ' ']
                        for drug in relevant_drugs:    
                            drugs_sum.loc[len(drugs_sum)] = [sample, gene, mutation, drug, row[drug]]


        drugs_sum.loc[drugs_sum["notes"] == "Polymorphism","drugs"] = "" 
        drugs_sum = drugs_sum.drop_duplicates().sort_values(by=['sample','notes'])
        
        self.excel_colors(drugs_sum)
        
        
#override report
    def qc_report(self, bam_path, depth_path, output_report):
        '''
        run unique hsv functions.        
        generate qc report considering the genes.
        Parameters
        ----------
        bam_path : str
            path to bam folder.
        depth_path : str
            path to depth folder.
        output_report : str
            path to qc report file.

        Returns
        -------
        None.

        '''
        self.cut_genes("alignment/")
        self.mutations()
        self.resist_poly_reports()
        self.resist_poly_summary()
        
        general_qc = pd.DataFrame(columns=['sample', 'mapped%','mapped_reads','total_reads','cov_bases',"chimeric_read_count"])
        ul23_qc = pd.DataFrame(columns=['sample','coverage%','coverage_CNS_5%', 'mean_depth','max_depth','min_depth'])
        ul30_qc = pd.DataFrame(columns=['sample','coverage%','coverage_CNS_5%', 'mean_depth','max_depth','min_depth'])
        ul42_qc = pd.DataFrame(columns=['sample','coverage%','coverage_CNS_5%', 'mean_depth','max_depth','min_depth'])
        
        for bam_file in os.listdir(bam_path):
                if "sorted" in bam_file and "bai" not in bam_file:
                    #general qc
                    sample = bam_file.split(".mapped")[0] + bam_file.split(".sorted")[1].split(".bam")[0]
                    total_reads = pysam.AlignmentFile(bam_path + bam_file.split(".mapped")[0]+".bam").count(until_eof=True)
                    mapped_reads, mapped_percentage, cov_bases, chimer_count = super().general_qc(bam_path+bam_file, total_reads)
            
                    #depth 
                    max_depth23, min_depth23,mean_depth23, cover23, cns5_cover23 = super().depth_qc(depth_path+sample+".txt", start = ul23_reg[0], end= ul23_reg[1])
                    max_depth30, min_depth30, mean_depth30, cover30, cns5_cover30 = super().depth_qc(depth_path+sample+".txt", start = ul30_reg[0], end= ul30_reg[1])
                    max_depth42, min_depth42,mean_depth42, cover42, cns5_cover42 = super().depth_qc(depth_path+sample+".txt", start = ul42_reg[0], end= ul42_reg[1])
                    
                    #add rows
                    general_qc.loc[len(general_qc.index)] = [sample, mapped_percentage, mapped_reads, total_reads, cov_bases, chimer_count]
                    ul23_qc.loc[len(ul23_qc.index)] = [sample, cover23, cns5_cover23, mean_depth23, max_depth23, min_depth23]
                    ul30_qc.loc[len(ul30_qc.index)] = [sample, cover30, cns5_cover30, mean_depth30, max_depth23, min_depth30]
                    ul42_qc.loc[len(ul42_qc.index)] = [sample, cover42, cns5_cover42, mean_depth42, max_depth42, min_depth42]
                    
        with pd.ExcelWriter('QC/QC_report.xlsx') as writer:  
            general_qc.to_excel(writer, sheet_name='general', index=False)
            ul23_qc.to_excel(writer, sheet_name='UL23', index=False)
            ul30_qc.to_excel(writer, sheet_name='UL30', index=False)
            ul42_qc.to_excel(writer, sheet_name='UL42', index=False)
                        
         